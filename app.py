import openpyxl

# wb = openpyxl.Workbook()

wb = openpyxl.load_workbook("transactions.xlsx")
# print(wb.sheetnames)

sheet = wb["Sheet1"]

for row in range(1, 10):
    cell = sheet.cell(row, 1)
    print(cell.value)
sheet.append([1, 2, 3])
wb.save("transactions1.xlsx")

# cell = sheet["a1"]

# column = sheet["a"]

# cells = sheet["a:c"]

# sheet["a1:c3"]

# sheet[1:3]

# sheet.append([1, 2, 3])


# print(column)

# for row in range(1, sheet.max_row+1):
#     for column in range(1, sheet.max_column+1):
#         cell = sheet.cell(row, column)
#         print(cell.value)
